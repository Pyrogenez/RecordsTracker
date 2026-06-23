"""Generate the analysis workbook from SQLite, and read user overrides back."""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .database import Database

log = logging.getLogger(__name__)

OVERRIDES_SHEET = "Overrides"
OVERRIDES_HEADERS = [
    "request_id",
    "is_closed",
    "first_real_reply_message_id",
    "notes",
    "updated_at",
]

REQUESTS_HEADERS = [
    "request_id", "rid", "status", "final_state", "request_type",
    "category", "department", "records_type", "description",
    "preferred_method", "requester_email",
    "submission_time", "first_auto_ack_time", "first_real_reply_time",
    "hours_to_first_reply",
    "first_seen_at", "last_scraped_at", "last_modified_at",
    "detail_url",
]
MESSAGES_HEADERS = [
    "request_id", "message_id", "sequence_num", "sent_at", "sender",
    "is_auto_ack", "subject", "body",
]
ATTACHMENTS_HEADERS = [
    "request_id", "attachment_id", "filename", "download_status",
    "file_size", "downloaded_at", "local_path", "error_message",
]
COMPLIANCE_HEADERS = [
    "issue_id", "request_id", "statute_section", "issue_type", "severity",
    "status", "description", "evidence", "ai_confidence", "identified_by",
    "model", "user_notes", "created_at", "updated_at",
]


def sync_overrides_from_excel(db: Database, excel_path: Path) -> int:
    """Read the Overrides sheet and write any user-edited rows back to SQLite.

    Returns the number of overrides applied. If the workbook does not exist yet,
    returns 0 silently. If it exists but is corrupted / locked / unreadable
    (e.g. partial OneDrive sync, open in Excel, crashed mid-write), we log a
    warning and return 0 so the scraper run can proceed. The workbook will be
    regenerated at the end of the run, which usually fixes the corruption."""
    if not excel_path.exists():
        return 0
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        log.warning(
            "Could not read existing Excel workbook at %s (%s: %s). "
            "Skipping override sync for this run — the workbook will be "
            "regenerated at the end of the run.",
            excel_path, type(e).__name__, e,
        )
        return 0
    if OVERRIDES_SHEET not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb[OVERRIDES_SHEET]
    headers: list[str] = []
    applied = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c) if c is not None else "" for c in row]
            continue
        record = dict(zip(headers, row))
        request_id = (record.get("request_id") or "").strip() if isinstance(record.get("request_id"), str) else None
        if not request_id:
            continue
        msg_id_raw = record.get("first_real_reply_message_id")
        msg_id: int | None
        if msg_id_raw in (None, ""):
            msg_id = None
        else:
            try:
                msg_id = int(msg_id_raw)
            except (TypeError, ValueError):
                msg_id = None
        notes = record.get("notes") or None
        if isinstance(notes, str):
            notes = notes.strip() or None
        is_closed = _parse_bool(record.get("is_closed"))
        # Skip writing a blank override if nothing is set
        if msg_id is None and not is_closed and not notes:
            continue
        db.upsert_override(request_id, msg_id, notes, is_closed=is_closed)
        applied += 1
    wb.close()
    return applied


def _parse_bool(value) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    s = str(value).strip().lower()
    return s in {"1", "true", "yes", "y", "closed", "x"}


def write_workbook(db: Database, excel_path: Path) -> None:
    """(Re)generate the workbook from SQLite. Preserves user overrides
    (which were already synced into SQLite before this is called)."""
    wb = Workbook()
    # Default sheet -> Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary(ws_summary, db)

    ws_req = wb.create_sheet("Requests")
    _write_requests(ws_req, db)

    ws_msgs = wb.create_sheet("Messages")
    _write_messages(ws_msgs, db)

    ws_att = wb.create_sheet("Attachments")
    _write_attachments(ws_att, db)

    ws_ov = wb.create_sheet(OVERRIDES_SHEET)
    _write_overrides(ws_ov, db)

    ws_comp = wb.create_sheet("Compliance Issues")
    _write_compliance(ws_comp, db)

    summaries = db.get_all_request_summaries()
    if summaries:
        ws_sum = wb.create_sheet("AI Summaries")
        _write_summaries(ws_sum, summaries)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    # Atomic write: save to a sibling temp file, then replace. If the process
    # is killed mid-save (Ctrl+C, OneDrive sync conflict, Excel locking the
    # file), we never leave a truncated .xlsx behind to poison the next run.
    tmp_path = excel_path.with_suffix(excel_path.suffix + ".tmp")
    try:
        wb.save(tmp_path)
        import os
        os.replace(tmp_path, excel_path)
    except Exception:
        # Clean up the partial temp file so it doesn't accumulate.
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except OSError:
            pass
        raise


def _write_summaries(ws, rows: list[dict]) -> None:
    headers = ["request_id", "summary", "model", "updated_at"]
    ws.append(headers)
    out_rows: list[list[Any]] = []
    for r in rows:
        row = [r.get(h) for h in headers]
        out_rows.append(row)
        ws.append(row)
    _style_header(ws, len(headers))
    _autosize(ws, headers, out_rows, cap=120)
    _add_table(ws, "SummariesTable", headers, len(out_rows))
    # Wrap long summary cells
    for i in range(2, len(out_rows) + 2):
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")


# ---- helpers ----

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
_OVERRIDE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.freeze_panes = "A2"


def _autosize(ws, headers: list[str], rows: list[list[Any]], cap: int = 60) -> None:
    for idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for r in rows:
            v = r[idx - 1]
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, cap)


def _add_table(ws, name: str, headers: list[str], n_rows: int) -> None:
    if n_rows == 0:
        return
    ref = f"A1:{get_column_letter(len(headers))}{n_rows + 1}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True,
    )
    ws.add_table(table)


def _write_requests(ws, db: Database) -> None:
    ws.append(REQUESTS_HEADERS)
    rows: list[list[Any]] = []
    for r in db.get_all_requests():
        row = [r.get(h) for h in REQUESTS_HEADERS]
        rows.append(row)
        ws.append(row)
    _style_header(ws, len(REQUESTS_HEADERS))
    _autosize(ws, REQUESTS_HEADERS, rows)
    _add_table(ws, "RequestsTable", REQUESTS_HEADERS, len(rows))


def _write_messages(ws, db: Database) -> None:
    ws.append(MESSAGES_HEADERS)
    rows: list[list[Any]] = []
    for m in db.get_all_messages():
        row = [m.get(h) for h in MESSAGES_HEADERS]
        rows.append(row)
        ws.append(row)
    _style_header(ws, len(MESSAGES_HEADERS))
    _autosize(ws, MESSAGES_HEADERS, rows, cap=80)
    _add_table(ws, "MessagesTable", MESSAGES_HEADERS, len(rows))


def _write_attachments(ws, db: Database) -> None:
    ws.append(ATTACHMENTS_HEADERS)
    rows: list[list[Any]] = []
    for a in db.get_all_attachments():
        row = [a.get(h) for h in ATTACHMENTS_HEADERS]
        rows.append(row)
        ws.append(row)
    _style_header(ws, len(ATTACHMENTS_HEADERS))
    _autosize(ws, ATTACHMENTS_HEADERS, rows, cap=80)
    _add_table(ws, "AttachmentsTable", ATTACHMENTS_HEADERS, len(rows))


def _write_compliance(ws, db: Database) -> None:
    ws.append(COMPLIANCE_HEADERS)
    rows: list[list[Any]] = []
    for iss in db.get_compliance_issues():
        row = [iss.get(h) for h in COMPLIANCE_HEADERS]
        rows.append(row)
        ws.append(row)
    _style_header(ws, len(COMPLIANCE_HEADERS))
    _autosize(ws, COMPLIANCE_HEADERS, rows, cap=80)
    _add_table(ws, "ComplianceTable", COMPLIANCE_HEADERS, len(rows))
    # Wrap long description cells for readability
    desc_col = COMPLIANCE_HEADERS.index("description") + 1
    evid_col = COMPLIANCE_HEADERS.index("evidence") + 1
    for i in range(2, len(rows) + 2):
        ws.cell(row=i, column=desc_col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=i, column=evid_col).alignment = Alignment(wrap_text=True, vertical="top")


def _write_overrides(ws, db: Database) -> None:
    ws.append(OVERRIDES_HEADERS)
    rows: list[list[Any]] = []
    existing = {o["request_id"]: o for o in db.get_all_overrides()}
    # Show every request so the user has a row to edit even if no override exists yet.
    for r in db.get_all_requests():
        rid = r["request_id"]
        ov = existing.get(rid, {})
        row = [
            rid,
            bool(ov.get("is_closed")),
            ov.get("first_real_reply_message_id"),
            ov.get("notes"),
            ov.get("updated_at"),
        ]
        rows.append(row)
        ws.append(row)
    _style_header(ws, len(OVERRIDES_HEADERS))
    _autosize(ws, OVERRIDES_HEADERS, rows, cap=60)
    _add_table(ws, "OverridesTable", OVERRIDES_HEADERS, len(rows))

    ws.cell(row=1, column=2).comment = _comment(
        "Set to TRUE (or 1, or 'closed') to tell the scraper to stop checking "
        "this request on incremental runs — regardless of what the portal says. "
        "Leave FALSE/blank to let the portal's status decide."
    )
    ws.cell(row=1, column=3).comment = _comment(
        "Override the auto-detected 'first real reply'. Paste the message_id "
        "(see Messages sheet) you consider the first substantive reply. "
        "Leave blank to use the default heuristic or AI classification."
    )
    ws.cell(row=1, column=4).comment = _comment(
        "Free-text notes saved with the override. Optional."
    )


def _comment(text: str):
    from openpyxl.comments import Comment
    return Comment(text, "records_tracker")


def _write_summary(ws, db: Database) -> None:
    ws["A1"] = "St. Pete Public Records Tracker"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(italic=True)

    requests = db.get_all_requests()
    messages = db.get_all_messages()
    attachments = db.get_all_attachments()

    completed = [r for r in requests if (r.get("final_state") or "").lower() == "completed"]
    in_progress = [r for r in requests if (r.get("final_state") or "").lower() == "in progress"]
    with_reply = [r for r in requests if r.get("hours_to_first_reply") is not None]
    if with_reply:
        avg_hours = sum(r["hours_to_first_reply"] for r in with_reply) / len(with_reply)
    else:
        avg_hours = None
    downloaded = [a for a in attachments if a.get("download_status") == "downloaded"]
    pending = [a for a in attachments if a.get("download_status") in ("pending", "failed")]

    counts = db.counts()
    rows = [
        ("Total requests", len(requests)),
        ("  Completed", len(completed)),
        ("  In progress", len(in_progress)),
        ("  Open (tracked on next run)", counts["open_requests"]),
        ("  Closed by user", counts["user_closed_requests"]),
        ("Total messages", len(messages)),
        ("Total attachments", len(attachments)),
        ("  Downloaded", len(downloaded)),
        ("  Pending / failed", len(pending)),
        ("  With extracted text", counts["attachments_with_text"]),
        ("Classified messages (AI)", counts["classified_messages"]),
        ("Summarized requests (AI)", counts["summarized_requests"]),
        ("Saved AI conversations", counts.get("conversations", 0)),
        ("Compliance issues (open)", counts.get("compliance_issues_open", 0)),
        ("Compliance issues (total)", counts.get("compliance_issues_total", 0)),
        (
            "Avg hours to first reply",
            round(avg_hours, 2) if avg_hours is not None else "—",
        ),
        ("Sample size for avg", len(with_reply)),
    ]
    for i, (label, value) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        ws.cell(row=i, column=1).font = Font(bold=label.startswith(" ") is False)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
