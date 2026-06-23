"""AI analysis layer for the records tracker.

Reads the SQLite database populated by run.py and enriches it with:
  * text extracted from downloaded attachments (PDF/DOCX/XLSX/TXT)
  * AI classifications of each STPETEFL message (auto_ack / status_update /
    substantive / other) — used to derive a more accurate
    'first real reply' time
  * AI-generated per-request summaries

Usage
-----
    python analyze.py extract      # extract text from attachments
    python analyze.py classify     # classify messages (uses Claude)
    python analyze.py summarize    # generate per-request summaries (uses Claude)
    python analyze.py all          # run extract -> classify -> summarize
    python analyze.py ask "Which requests mention police body camera footage?"
    python analyze.py ask --request P121302-042026 "What did they send?"
    python analyze.py stats        # quick counts of what's been analyzed

The Anthropic API key is read from the ANTHROPIC_API_KEY environment variable,
or from config.json under the key "anthropic_api_key".

Calls to Claude are cached in SQLite — re-running `classify` or `summarize`
is cheap (only new/changed items hit the API).
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from records_tracker.config import PROJECT_ROOT, load_config, project_paths
from records_tracker.database import Database, is_support_sender
from records_tracker.excel_export import write_workbook

log = logging.getLogger("analyze")

# Model choices: use Haiku for classification (cheap, high volume), Sonnet for
# summaries + ad-hoc Q&A (better reasoning, lower volume).
MODEL_CLASSIFY = "claude-haiku-4-5-20251001"
MODEL_SUMMARIZE = "claude-sonnet-4-6"
MODEL_ASK = "claude-sonnet-4-6"

# Max chars of attachment text to include in a single prompt before truncating.
ATTACHMENT_SNIPPET_LIMIT = 6000
# Max chars per attachment when packed into a multi-attachment prompt.
ATTACHMENT_PACKED_LIMIT = 2000


# =============================================================================
# Setup + small helpers
# =============================================================================

def setup_logging(log_dir: Path) -> None:
    log_dir.mkdir(parents=True, exist_ok=True)
    logfile = log_dir / f"{datetime.now().strftime('%Y-%m-%d')}-analyze.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(logfile, encoding="utf-8"),
        ],
    )


def _get_api_key() -> str | None:
    key = os.environ.get("ANTHROPIC_API_KEY")
    if key:
        return key
    # Resolve against the project root, not the current working directory.
    cfg_path = PROJECT_ROOT / "config.json"
    if cfg_path.exists():
        try:
            raw = json.loads(cfg_path.read_text())
            k = raw.get("anthropic_api_key")
            if k and isinstance(k, str) and k.strip():
                return k.strip()
        except Exception:
            pass
    return None


def _require_anthropic_client():
    """Lazily import and construct the Anthropic client. Exits with a friendly
    message if the key is missing or the package isn't installed."""
    key = _get_api_key()
    if not key:
        print(
            "ERROR: no Anthropic API key found.\n"
            "  Set the ANTHROPIC_API_KEY environment variable, or add\n"
            '    "anthropic_api_key": "sk-ant-..."\n'
            "  to config.json.",
            file=sys.stderr,
        )
        sys.exit(2)
    try:
        import anthropic
    except ImportError:
        print(
            "ERROR: the 'anthropic' package is not installed.\n"
            "  Run:  python -m pip install -r requirements.txt",
            file=sys.stderr,
        )
        sys.exit(2)
    return anthropic.Anthropic(api_key=key, max_retries=4)


# =============================================================================
# extract — pull text out of downloaded attachment files
# =============================================================================

def extract_text_from_file(path: Path) -> tuple[str | None, str | None]:
    """Return (text, error). Exactly one is non-None."""
    if not path.exists():
        return None, f"file missing: {path}"
    suffix = path.suffix.lower()
    try:
        if suffix == ".pdf":
            return _extract_pdf(path), None
        if suffix in {".docx"}:
            return _extract_docx(path), None
        if suffix in {".xlsx", ".xlsm"}:
            return _extract_xlsx(path), None
        if suffix in {".txt", ".csv", ".log", ".md", ".json", ".xml", ".html", ".htm"}:
            return path.read_text(encoding="utf-8", errors="replace"), None
        return None, f"unsupported type: {suffix}"
    except Exception as e:
        return None, f"{type(e).__name__}: {e}"


def _extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as e:
        raise RuntimeError("pypdf not installed; run pip install -r requirements.txt") from e
    reader = PdfReader(str(path))
    parts: list[str] = []
    for i, page in enumerate(reader.pages):
        try:
            t = page.extract_text() or ""
        except Exception:
            t = ""
        if t.strip():
            parts.append(f"--- page {i + 1} ---\n{t}")
    return "\n\n".join(parts)


def _extract_docx(path: Path) -> str:
    try:
        from docx import Document
    except ImportError as e:
        raise RuntimeError("python-docx not installed; run pip install -r requirements.txt") from e
    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def _extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        parts.append(f"=== sheet: {sheet} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            parts.append("\t".join(cells))
    wb.close()
    return "\n".join(parts)


def command_extract(db: Database, limit: int | None = None) -> int:
    pending = db.attachments_needing_text()
    if limit:
        pending = pending[:limit]
    if not pending:
        log.info("No attachments need text extraction.")
        return 0
    log.info("Extracting text from %d attachment(s)...", len(pending))
    ok = fail = 0
    for a in pending:
        local = a.get("local_path")
        if not local:
            db.upsert_attachment_text(a["attachment_id"], None, "no local_path")
            fail += 1
            continue
        text, err = extract_text_from_file(Path(local))
        if err:
            log.warning("  [%s] %s — %s", a["request_id"], a["filename"], err)
            db.upsert_attachment_text(a["attachment_id"], None, err)
            fail += 1
        else:
            wc = len((text or "").split())
            log.info("  [%s] %s — %d words", a["request_id"], a["filename"], wc)
            db.upsert_attachment_text(a["attachment_id"], text, None)
            ok += 1
    log.info("Extraction complete. ok=%d failed=%d", ok, fail)
    return 0


# =============================================================================
# classify — label STPETEFL messages as auto_ack/status_update/substantive
# =============================================================================

_CLASSIFY_SYSTEM = (
    "You classify replies from a public-records support team at the St. Petersburg FL "
    "GovQA portal. Messages come from 'STPETEFL Support'. Classify each message as "
    "one of:\n"
    "  - auto_ack: the automatic 'we received your request' confirmation. "
    "Usually generic boilerplate about the request being logged.\n"
    "  - status_update: a short note that work is ongoing / extended / "
    "assigned / needs clarification. No records actually delivered.\n"
    "  - substantive: the first real response with records, an estimate, a "
    "determination, or a definitive answer to the request.\n"
    "  - other: anything that doesn't fit the above (e.g. closing note, "
    "survey request).\n\n"
    "Return ONLY a JSON object like:\n"
    '  {\"classification\": \"substantive\", \"confidence\": 0.9, '
    '\"reasoning\": \"contains delivered records and a final determination\"}\n'
    "No markdown, no code fences."
)


def _classify_one(client, message: dict) -> dict:
    prompt = (
        f"Request: {message['request_id']}\n"
        f"Sender: {message['sender']}\n"
        f"Sent at: {message['sent_at']}\n"
        f"Subject: {message.get('subject') or ''}\n"
        f"Body:\n{(message.get('body') or '')[:4000]}\n"
    )
    resp = client.messages.create(
        model=MODEL_CLASSIFY,
        max_tokens=400,
        system=_CLASSIFY_SYSTEM,
        messages=[{"role": "user", "content": prompt}],
    )
    text = "".join(b.text for b in resp.content if getattr(b, "type", None) == "text")
    parsed = _safe_json(text) or {}
    classification = parsed.get("classification") or "other"
    if classification not in {"auto_ack", "status_update", "substantive", "other"}:
        classification = "other"
    return {
        "classification": classification,
        "confidence": parsed.get("confidence"),
        "reasoning": (parsed.get("reasoning") or "")[:500],
    }


def _safe_json(text: str) -> dict | None:
    text = (text or "").strip()
    if text.startswith("```"):
        # strip code fences
        text = text.strip("`")
        if text.lower().startswith("json"):
            text = text[4:]
    # find first { and last }
    i = text.find("{")
    j = text.rfind("}")
    if i == -1 or j == -1 or j < i:
        return None
    try:
        return json.loads(text[i:j + 1])
    except Exception:
        return None


def command_classify(db: Database, limit: int | None = None,
                     only_support: bool = True) -> int:
    pending = db.messages_needing_classification()
    if only_support:
        pending = [m for m in pending if is_support_sender(m["sender"])]
    if limit:
        pending = pending[:limit]
    if not pending:
        log.info("No messages need classification.")
        return 0
    client = _require_anthropic_client()
    log.info("Classifying %d message(s) with %s...", len(pending), MODEL_CLASSIFY)
    for i, m in enumerate(pending, 1):
        try:
            res = _classify_one(client, m)
            db.upsert_message_classification(
                m["message_id"],
                res["classification"],
                res["confidence"],
                res["reasoning"],
                MODEL_CLASSIFY,
            )
            log.info("  [%d/%d] msg=%d req=%s -> %s",
                     i, len(pending), m["message_id"], m["request_id"],
                     res["classification"])
        except Exception as e:  # noqa: BLE001
            log.error("  [%d/%d] msg=%d FAILED: %s", i, len(pending), m["message_id"], e)
    # Recompute first-reply times for affected requests so downstream stats
    # reflect the new classifications.
    affected = {m["request_id"] for m in pending}
    with db.transaction():
        for rid in affected:
            db.recompute_first_reply(rid)
    log.info("Classification complete. Recomputed first-reply for %d request(s).",
             len(affected))
    return 0


# =============================================================================
# summarize — per-request synthesis across messages + attachment text
# =============================================================================

_SUMMARIZE_SYSTEM = (
    "You summarize public records requests from the St. Petersburg FL portal. "
    "Write a clear, factual summary (1-2 short paragraphs) covering:\n"
    "  - what was requested\n"
    "  - what the city has provided so far (if anything)\n"
    "  - the current status and any outstanding items\n"
    "  - any notable dates or deadlines\n"
    "Be concise. Don't editorialize. If the records delivered contain "
    "specific figures, names, or incident details relevant to the request, "
    "mention them briefly. If the request is still pending, say so."
)


def _pack_request_context(db: Database, request_id: str) -> str:
    r = db.get_request(request_id)
    if not r:
        return ""
    messages = db.get_messages_for_request(request_id)
    attachments = db.get_attachments_for_request(request_id)

    parts: list[str] = []
    parts.append(f"Request ID: {r['request_id']}")
    parts.append(f"Status: {r.get('status')} / final_state: {r.get('final_state')}")
    parts.append(f"Type: {r.get('request_type')}  Dept: {r.get('department')}")
    parts.append(f"Submitted: {r.get('submission_time')}")
    parts.append(f"First real reply: {r.get('first_real_reply_time')}")
    parts.append(f"Hours to first reply: {r.get('hours_to_first_reply')}")
    parts.append("")
    parts.append(f"Description: {r.get('description') or '(none)'}")
    parts.append("")
    parts.append("--- Messages ---")
    for m in messages:
        body = (m.get("body") or "")[:1500]
        parts.append(f"[{m['sent_at']}] {m['sender']} — {m.get('subject') or ''}\n{body}")
    if attachments:
        parts.append("")
        parts.append("--- Attachments (with extracted text) ---")
        for a in attachments:
            parts.append(f"• {a['filename']}  ({a['download_status']})")
            txt = db.get_attachment_text(a["attachment_id"])
            if txt and txt.get("extracted_text"):
                snippet = txt["extracted_text"][:ATTACHMENT_PACKED_LIMIT]
                parts.append(snippet)
    return "\n".join(parts)


def _request_changed_since_summary(last_modified: str | None,
                                   summary_updated: str | None) -> bool:
    """True if a request looks newer than its existing summary.

    last_modified is a naive portal-local (Eastern) message time; summary
    updated_at is UTC. A raw string comparison (the previous logic) is wrong
    because of that offset and the '+00:00' suffix, and could silently leave a
    stale summary. We parse both, normalize to naive UTC, and add a slack window
    to the naive Eastern value so we err toward refreshing rather than skipping.
    """
    if not last_modified or not summary_updated:
        return True
    try:
        from dateutil import parser as _p
        lm = _p.parse(last_modified)
        su = _p.parse(summary_updated)
        lm_naive = lm.tzinfo is None
        if not lm_naive:
            lm = lm.astimezone(timezone.utc).replace(tzinfo=None)
        if su.tzinfo is not None:
            su = su.astimezone(timezone.utc).replace(tzinfo=None)
        slack = timedelta(hours=6) if lm_naive else timedelta(0)
        return (lm + slack) > su
    except Exception:
        return True


def command_summarize(db: Database, force: bool = False,
                      limit: int | None = None,
                      only_request: str | None = None) -> int:
    all_requests = db.get_all_requests()
    if only_request:
        all_requests = [r for r in all_requests if r["request_id"] == only_request]
    todo: list[dict] = []
    for r in all_requests:
        existing = db.get_request_summary(r["request_id"])
        if existing and not force:
            if not _request_changed_since_summary(
                    r.get("last_modified_at"), existing.get("updated_at")):
                continue
        todo.append(r)
    if limit:
        todo = todo[:limit]
    if not todo:
        log.info("No requests need summarization. (Use --force to redo existing.)")
        return 0
    client = _require_anthropic_client()
    log.info("Summarizing %d request(s) with %s...", len(todo), MODEL_SUMMARIZE)
    for i, r in enumerate(todo, 1):
        context = _pack_request_context(db, r["request_id"])
        if not context:
            continue
        try:
            resp = client.messages.create(
                model=MODEL_SUMMARIZE,
                max_tokens=600,
                system=_SUMMARIZE_SYSTEM,
                messages=[{"role": "user", "content": context[:40000]}],
            )
            text = "".join(b.text for b in resp.content if getattr(b, "type", None) == "text").strip()
            if text:
                db.upsert_request_summary(r["request_id"], text, MODEL_SUMMARIZE)
                log.info("  [%d/%d] %s — summarized (%d chars)",
                         i, len(todo), r["request_id"], len(text))
        except Exception as e:  # noqa: BLE001
            log.error("  [%d/%d] %s FAILED: %s", i, len(todo), r["request_id"], e)
    log.info("Summarization complete.")
    return 0


# =============================================================================
# ask — natural language Q&A against the whole corpus (or one request)
# =============================================================================

def _build_corpus(db: Database, only_request: str | None) -> str:
    if only_request:
        return _pack_request_context(db, only_request)

    # For whole-corpus questions, pack lightweight per-request digests:
    # request metadata + short summary if one exists + attachment filenames.
    parts: list[str] = []
    summaries = {s["request_id"]: s for s in db.get_all_request_summaries()}
    for r in db.get_all_requests():
        rid = r["request_id"]
        block = [
            f"### {rid}",
            f"type={r.get('request_type')}, dept={r.get('department')}, "
            f"status={r.get('status')}/{r.get('final_state')}",
            f"submitted={r.get('submission_time')}, "
            f"first_reply={r.get('first_real_reply_time')}, "
            f"hours_to_first_reply={r.get('hours_to_first_reply')}",
        ]
        desc = (r.get("description") or "").strip()
        if desc:
            block.append(f"description: {desc[:500]}")
        s = summaries.get(rid)
        if s:
            block.append(f"summary: {s['summary'][:1000]}")
        atts = [a for a in db.get_all_attachments() if a["request_id"] == rid]
        if atts:
            block.append("attachments: " + ", ".join(a["filename"] for a in atts[:20]))
        parts.append("\n".join(block))
    return "\n\n".join(parts)


_ASK_SYSTEM = (
    "You answer questions about a dataset of public records requests made to "
    "the City of St. Petersburg, FL. Use ONLY the provided context. If the "
    "context doesn't contain the answer, say so. Cite specific request IDs "
    "(like P121302-042026) when referring to individual requests. Be concise "
    "and direct. If the user asks for a list or comparison, format clearly."
)


def command_ask(db: Database, question: str, only_request: str | None = None) -> int:
    corpus = _build_corpus(db, only_request)
    if not corpus.strip():
        print("No data to answer against. Run run.py first.", file=sys.stderr)
        return 1
    client = _require_anthropic_client()
    # Truncate corpus if extremely large — keep well under context limits.
    max_chars = 180_000
    if len(corpus) > max_chars:
        corpus = corpus[:max_chars] + "\n\n[...truncated...]"
    user_msg = f"Context:\n{corpus}\n\n---\n\nQuestion: {question}"
    resp = client.messages.create(
        model=MODEL_ASK,
        max_tokens=1500,
        system=_ASK_SYSTEM,
        messages=[{"role": "user", "content": user_msg}],
    )
    text = "".join(b.text for b in resp.content if getattr(b, "type", None) == "text")
    print(text)
    return 0


# =============================================================================
# stats + all
# =============================================================================

def command_stats(db: Database) -> int:
    c = db.counts()
    print(f"Requests:            {c['total_requests']}")
    print(f"  Open:              {c['open_requests']}")
    print(f"  User-closed:       {c['user_closed_requests']}")
    print(f"Messages:            {c['total_messages']}")
    print(f"  Classified:        {c['classified_messages']}")
    print(f"Attachments:         {c['total_attachments']}")
    print(f"  Downloaded:        {c['downloaded_attachments']}")
    print(f"  Text extracted:    {c['attachments_with_text']}")
    print(f"Request summaries:   {c['summarized_requests']}")
    return 0


def command_all(db: Database, excel_path: Path, limit: int | None = None,
                assume_yes: bool = False) -> int:
    to_extract = len(db.attachments_needing_text())
    to_classify = len([m for m in db.messages_needing_classification()
                       if is_support_sender(m["sender"])])
    n_requests = len(db.get_all_requests())
    print("Pending AI work:")
    print(f"  attachments to extract text from: {to_extract}  (local, free)")
    print(f"  messages to classify (Claude):    {to_classify}")
    print(f"  requests to (re)summarize (Claude): up to {n_requests}")
    # classify + summarize call the paid API; guard against accidental spend,
    # especially when invoked non-interactively (e.g. from the web UI).
    if not assume_yes and limit is None:
        if sys.stdin.isatty():
            resp = input(
                "This calls the Anthropic API and may cost money. Proceed? [y/N] "
            ).strip().lower()
            if resp not in ("y", "yes"):
                print("Aborted.")
                return 0
        else:
            print("Refusing to run non-interactively without --yes (would call "
                  "the paid API).\nRe-run with:  python analyze.py all --yes  "
                  "(optionally --limit N)", file=sys.stderr)
            return 1
    command_extract(db, limit=limit)
    command_classify(db, limit=limit)
    command_summarize(db, limit=limit)
    # Regenerate Excel so the AI Summaries sheet + updated first-reply times show up.
    try:
        write_workbook(db, excel_path)
        log.info("Excel workbook regenerated at %s", excel_path)
    except Exception:
        log.exception("Failed to regenerate Excel workbook")
    return 0


# =============================================================================
# CLI
# =============================================================================

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="AI analysis over the records database")
    sub = p.add_subparsers(dest="command", required=True)

    ex = sub.add_parser("extract", help="Extract text from downloaded attachments")
    ex.add_argument("--limit", type=int, help="Process at most N attachments")

    cl = sub.add_parser("classify", help="Classify messages (Claude)")
    cl.add_argument("--limit", type=int, help="Classify at most N messages")
    cl.add_argument("--all-senders", action="store_true",
                    help="Classify every sender, not just STPETEFL Support")

    sm = sub.add_parser("summarize", help="Summarize requests (Claude)")
    sm.add_argument("--force", action="store_true",
                    help="Redo summaries even if one already exists")
    sm.add_argument("--limit", type=int, help="Summarize at most N requests")
    sm.add_argument("--request", metavar="REQUEST_ID",
                    help="Restrict to one request ID")

    al = sub.add_parser("all", help="Run extract -> classify -> summarize")
    al.add_argument("--limit", type=int, help="Limit each stage to N items")
    al.add_argument("--yes", action="store_true",
                    help="Skip the confirmation prompt (required when run "
                         "non-interactively, e.g. from the web UI)")

    ask = sub.add_parser("ask", help="Ask a natural-language question")
    ask.add_argument("question", help="The question to ask")
    ask.add_argument("--request", metavar="REQUEST_ID",
                     help="Restrict context to a single request")

    sub.add_parser("stats", help="Print DB counts and analysis coverage")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    paths = project_paths()
    setup_logging(paths["logs"])

    if not paths["database"].exists():
        print("No database yet. Run `python run.py --full` first.", file=sys.stderr)
        return 1
    db = Database(paths["database"])
    try:
        if args.command == "extract":
            return command_extract(db, limit=args.limit)
        if args.command == "classify":
            return command_classify(
                db,
                limit=args.limit,
                only_support=not args.all_senders,
            )
        if args.command == "summarize":
            return command_summarize(
                db,
                force=args.force,
                limit=args.limit,
                only_request=args.request,
            )
        if args.command == "all":
            return command_all(db, paths["excel"], limit=args.limit,
                               assume_yes=args.yes)
        if args.command == "ask":
            return command_ask(db, args.question, only_request=args.request)
        if args.command == "stats":
            return command_stats(db)
    finally:
        db.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
